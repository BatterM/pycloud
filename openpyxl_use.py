#!/usr/bin/env python3
#
#date:2019-8-22
#author:BatterM
#usage:use openpyxl

import openpyxl
# 1、打开已存在的文件，修改A1位置信息为good
# book=openpyxl.load_workbook('c:\\Users\\WL\\Desktop\\open.xlsx')
# sheet=book.active
# sheet['A1']='good'
# book.save('c:\\Users\\WL\\Desktop\\open.xlsx')
#
# 2、新建一个excel表格
# book=openpyxl.Workbook()
# book.save('openpyxl_use.xlsx')
#
# 3、指定表打开，可以迭代写入列表、元组
# book=openpyxl.load_workbook('openpyxl_use.xlsx')
# sheet=book['aaa']
# sheet.append([1,5,'6,8'])
# sheet=book['bbb']
# sheet.append((66,234,'234gsg'))
# book.save('openpyxl_use.xlsx')
a=65
b=chr(a)
print(b)

a='A'
b=ord(a)
print(b)

